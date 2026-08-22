"""Builds downloadable CSV / Excel / PDF reports for the "Export Data" feature.

Each entry in EXPORT_SOURCES maps a frontend page to the existing service
function that already powers that page's table, plus the curated set of
columns worth putting in a report (raw DB rows carry a lot of internal
JSON blobs that don't belong in an export).
"""

import csv
import io
import json
from datetime import date, datetime, timedelta

import fitz  # PyMuPDF - already used elsewhere for statement parsing
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from ..models.organization import Organization
from .accounts_service import get_paginated_accounts
from .transaction_service import get_paginated_transactions, query_audit_logs

# Hard cap so a single export can't try to pull an unbounded number of rows
# into memory / onto PDF pages.
EXPORT_ROW_LIMIT = 5000

# Same field-name -> human label mapping the Audit Log page uses for its own
# expanded "changes" view, so exports read the same way the UI does.
CHANGE_FIELD_LABELS = {
    "category": "Category",
    "narration": "Narration",
    "counterparty": "Counterparty",
    "account_number": "Account Number",
    "account_holder_name": "Account Holder Name",
    "txn_date": "Transaction Date",
    "mode": "Mode",
    "ref_number": "Reference ID",
    "amount": "Amount",
    "txn_type": "Transaction Type",
}

# Column tuples are (key, label, default) - "default" marks the columns
# that are already visible on that page's own table, so the export dialog
# can preselect exactly what the org is already looking at. Columns with
# default=False are still exportable, just opt-in (e.g. "Mode" isn't a
# column on the Transactions table itself).
#
# Accounts splits "last updated" into its two real timestamps instead of
# collapsing them into one ambiguous date.
EXPORT_SOURCES = {
    "transactions": {
        "label": "All Transactions",
        "filename": "transactions",
        "columns": [
            ("txn_date", "Date", True),
            ("bank_name", "Bank", True),
            ("account_holder_name", "Account Holder", True),
            ("account_number", "Account Number", True),
            ("counterparty", "Counterparty", True),
            ("txn_type", "Type", True),
            ("amount", "Amount", True),
            ("category", "Category", True),
            ("mode", "Mode", False),
            ("narration", "Narration", True),
            ("source", "Source", True),
            ("ref_number", "Reference Number", False),
        ],
    },
    "accounts": {
        "label": "All Accounts",
        "filename": "accounts",
        "columns": [
            ("account_holder_name", "Account Holder", True),
            ("account_number", "Account Number", True),
            ("bank_name", "Bank", True),
            ("category", "Category", True),
            ("account_type", "Type", True),
            ("statement_balance", "Balance - Statement", True),
            ("statement_updated_at", "Statement Last Updated", True),
            ("calculated_balance", "Balance - Calculated", True),
            ("calculated_updated_at", "Calculated Last Updated", True),
            ("delta", "Delta", True),
            ("source", "Source", False),
        ],
    },
    "audit-log": {
        "label": "Audit Logs",
        "filename": "audit-log",
        "columns": [
            ("changed_by", "Operator", True),
            ("created_at", "When", True),
            ("txn_narration", "Narration", True),
            ("changes", "Changes", True),
            ("reason", "Reason", True),
            ("ip_address", "IP Address", True),
            ("txn_amount", "Amount", False),
            ("txn_id", "Transaction ID", False),
        ],
    },
}

# Filter key -> human label, matching each page's own filter panel wording.
# Used to build the "Filters applied" summary shown in the export dialog and
# printed on the PDF cover page - the same filters dict a page already sends
# to its own /query endpoint is sent here unchanged and described from it.
TRANSACTION_FILTER_LABELS = {
    "search": "Search",
    "bank": "Bank",
    "account": "Account Number",
    "txnType": "Transaction Type",
    "mode": "Mode",
    "category": "Category",
    "currency": "Currency",
    "accountHolderName": "Account Holder",
    "accountType": "Account Type",
    "status": "Status",
    "entity": "Entity",
    "individualAccount": "Individual Account",
}
ACCOUNT_FILTER_LABELS = {
    "search": "Search",
    "account": "Account Number",
    "bank": "Bank",
    "accountType": "Account Type",
    "category": "Category",
    "accountHolderName": "Account Holder",
    "individualAccount": "Individual Account",
}
AUDIT_LOG_FILTER_LABELS = {
    "search": "Search",
    "changed_by": "Operator",
}
FILTER_LABELS_BY_SOURCE = {
    "transactions": TRANSACTION_FILTER_LABELS,
    "accounts": ACCOUNT_FILTER_LABELS,
    "audit-log": AUDIT_LOG_FILTER_LABELS,
}


def get_export_sources() -> list[dict]:
    """Metadata for the frontend's page picker + column checklist (no data
    fetching) - "default" tells the dialog which columns to preselect."""
    return [
        {
            "key": key,
            "label": meta["label"],
            "columns": [
                {"key": col_key, "label": col_label, "default": is_default}
                for col_key, col_label, is_default in meta["columns"]
            ],
        }
        for key, meta in EXPORT_SOURCES.items()
    ]


def _resolve_columns(source: str, requested_keys: list[str] | None) -> list[tuple[str, str]]:
    """Validates the requested column keys against the source's real columns
    (preserving canonical order) and falls back to the page's own default
    columns when nothing valid was requested."""
    all_columns = EXPORT_SOURCES[source]["columns"]
    default_columns = [(key, label) for key, label, is_default in all_columns if is_default]

    if not requested_keys:
        return default_columns

    requested_set = set(requested_keys)
    resolved = [(key, label) for key, label, _ in all_columns if key in requested_set]
    return resolved or default_columns


def _active_filter_values(value) -> list[str]:
    if not value or value == "all":
        return []
    values = value if isinstance(value, list) else [value]
    return [str(v).strip() for v in values if v and str(v).strip().lower() != "all"]


def _describe_filters(source: str, filters: dict | None) -> tuple[str | None, list[str]]:
    """Returns (date_line, filter_lines) - the report-metadata date/range
    shown on its own row, and the list of other active filters (skipping
    anything left at "all"/empty) for the "Filters applied" block."""
    filters = filters or {}
    labels = FILTER_LABELS_BY_SOURCE.get(source, {})
    filter_lines = []

    for key, label in labels.items():
        values = _active_filter_values(filters.get(key))
        if values:
            filter_lines.append(f"{label}: {', '.join(values)}")

    date_line = None

    if source == "transactions":
        min_amt, max_amt = filters.get("minAmount"), filters.get("maxAmount")
        if min_amt not in (None, "") or max_amt not in (None, ""):
            low = min_amt if min_amt not in (None, "") else "0"
            high = max_amt if max_amt not in (None, "") else "no limit"
            filter_lines.append(f"Amount: {low} - {high}")

        tab = filters.get("tab")
        if tab and tab != "transactions":
            filter_lines.append(f"Tab: {tab.replace('-', ' ').title()}")

        date_range = filters.get("dateRange") or {}
        start, end = date_range.get("startDate"), date_range.get("endDate")
        if start or end:
            date_line = f"{_format_date(start) or 'any'} - {_format_date(end) or 'any'}"

    elif source == "audit-log":
        start, end = filters.get("start_date"), filters.get("end_date")
        if start or end:
            date_line = f"{_format_date(start) or 'any'} - {_format_date(end) or 'any'}"

    elif source == "accounts":
        as_of = filters.get("date")
        if as_of:
            date_line = _format_date(as_of) or None

    return date_line, filter_lines


DEFAULT_DATE_RANGE_DAYS = 30


def _default_date_bounds(days: int = DEFAULT_DATE_RANGE_DAYS) -> tuple[str, str]:
    end = datetime.utcnow().date()
    start = end - timedelta(days=days)
    return start.isoformat(), end.isoformat()


def _with_default_date_range(source: str, filters: dict | None) -> dict:
    """Transactions/Audit Log pages always show a trailing date window by
    default (even before the org explicitly filters anything) - the export
    should match that instead of silently dumping the entire unbounded
    history just because no filter context happened to be published."""
    filters = dict(filters or {})

    if source == "transactions":
        date_range = dict(filters.get("dateRange") or {})
        if not date_range.get("startDate") and not date_range.get("endDate"):
            start, end = _default_date_bounds()
            date_range = {"startDate": start, "endDate": end}
        filters["dateRange"] = date_range

    elif source == "audit-log":
        if not filters.get("start_date") and not filters.get("end_date"):
            start, end = _default_date_bounds()
            filters["start_date"] = start
            filters["end_date"] = end

    elif source == "accounts":
        if not filters.get("date"):
            filters["date"] = datetime.utcnow().date().isoformat()

    return filters


def _fetch_rows(db: Session, org_id: int, source: str, filters: dict | None) -> list[dict]:
    filters = filters or {}

    if source == "transactions":
        result = get_paginated_transactions(db, org_id, filters, page=1, page_size=EXPORT_ROW_LIMIT)
        return result["data"]

    if source == "accounts":
        result = get_paginated_accounts(db, org_id, filters, page=1, page_size=1000)
        return result["accounts"]

    if source == "audit-log":
        result = query_audit_logs(
            db,
            org_id,
            page=1,
            page_size=EXPORT_ROW_LIMIT,
            search=filters.get("search") or None,
            changed_by=filters.get("changed_by") or None,
            start_date=filters.get("start_date") or None,
            end_date=filters.get("end_date") or None,
        )
        return result["logs"]

    raise ValueError(f"Unknown export source: {source}")


# --- value formatting -------------------------------------------------------
#
# Root cause of the original "date column is blank" bug: raw ISO timestamps
# ("2026-07-13T05:30:00+05:30") are a single unbroken token with no spaces.
# PyMuPDF's insert_textbox() word-wraps on spaces only, and silently draws
# nothing when a single word can't fit the cell - so every date-shaped
# column vanished in the PDF even though the underlying data was fine.
# Formatting dates into short, space-separated text ("13 Jul 2026") fixes
# both the PDF rendering and makes CSV/Excel more readable.

def _coerce_datetime(value):
    if isinstance(value, (datetime, date)):
        return value
    if isinstance(value, str) and value:
        try:
            return datetime.fromisoformat(value)
        except ValueError:
            return None
    return None


def _format_date(value) -> str:
    parsed = _coerce_datetime(value)
    return parsed.strftime("%d %b %Y") if parsed else ""


def _format_datetime(value) -> str:
    parsed = _coerce_datetime(value)
    return parsed.strftime("%d %b %Y, %H:%M") if parsed else ""


def _change_field_label(field: str) -> str:
    return CHANGE_FIELD_LABELS.get(field, field.replace("_", " ").title())


def _change_value_text(value) -> str:
    if value is None or value == "":
        return "(empty)"
    return str(value)


def _format_changes_lines(changes) -> list[str]:
    """One "Field: old -> new" line per changed field, matching the Audit Log
    page's own expanded-row wording (fieldLabels + old/new chips)."""
    if not isinstance(changes, dict) or not changes:
        return []

    lines = []
    for field, delta in changes.items():
        if not isinstance(delta, dict):
            continue
        old_text = _change_value_text(delta.get("old"))
        new_text = _change_value_text(delta.get("new"))
        lines.append(f"{_change_field_label(field)}: {old_text} -> {new_text}")
    return lines


# Formatters applied before the generic string-coercion / fallback logic.
# PDF gets its own override for txn_date: dropping the time keeps it to one
# short line so it reliably fits the narrow report columns.
COLUMN_FORMATTERS = {
    "txn_date": _format_datetime,
    "created_at": _format_datetime,
    "statement_updated_at": _format_date,
    "calculated_updated_at": _format_date,
}
PDF_COLUMN_FORMATTERS = {**COLUMN_FORMATTERS, "txn_date": _format_date}

# Fields worth a friendlier placeholder than a bare "-" when empty, matching
# what the actual pages show for the same blank fields.
FIELD_FALLBACKS = {
    "narration": "No narration available",
    "category": "Others",
}
DEFAULT_FALLBACK = "-"


def _cell_value(row: dict, key: str, formatters: dict | None = None, multiline: bool = False) -> str:
    if key == "changes":
        lines = _format_changes_lines(row.get(key))
        if not lines:
            return DEFAULT_FALLBACK
        return "\n".join(lines) if multiline else "; ".join(lines)

    value = row.get(key)
    formatter = (formatters or COLUMN_FORMATTERS).get(key)

    if formatter:
        formatted = formatter(value)
        return formatted or FIELD_FALLBACKS.get(key, DEFAULT_FALLBACK)

    if value is None or value == "":
        return FIELD_FALLBACKS.get(key, DEFAULT_FALLBACK)

    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)

    return str(value)


def build_export(
    db: Session,
    org: Organization,
    source: str,
    export_format: str,
    columns: list[str] | None = None,
    filters: dict | None = None,
) -> tuple[bytes, str, str]:
    """Returns (file_bytes, filename, content_type)."""
    if source not in EXPORT_SOURCES:
        raise ValueError(f"Unknown export source: {source}")

    meta = EXPORT_SOURCES[source]
    effective_filters = _with_default_date_range(source, filters)
    rows = _fetch_rows(db, org.id, source, effective_filters)
    resolved_columns = _resolve_columns(source, columns)
    timestamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    base_filename = f"{meta['filename']}-{timestamp}"

    if export_format == "csv":
        return _to_csv(rows, resolved_columns), f"{base_filename}.csv", "text/csv"

    if export_format == "xlsx":
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return _to_xlsx(rows, resolved_columns, meta["label"]), f"{base_filename}.xlsx", content_type

    if export_format == "pdf":
        date_line, filter_lines = _describe_filters(source, effective_filters)
        date_label = "As Of" if source == "accounts" else "Date Range"
        pdf_bytes = _build_pdf(source, rows, resolved_columns, meta["label"], org, date_label, date_line, filter_lines)
        return pdf_bytes, f"{base_filename}.pdf", "application/pdf"

    raise ValueError(f"Unsupported export format: {export_format}")


def _to_csv(rows: list[dict], columns: list[tuple[str, str]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([label for _, label in columns])
    for row in rows:
        writer.writerow([_cell_value(row, key) for key, _ in columns])
    # UTF-8 BOM so Excel opens accented / rupee-symbol text correctly.
    return buffer.getvalue().encode("utf-8-sig")


def _to_xlsx(rows: list[dict], columns: list[tuple[str, str]], sheet_title: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = (sheet_title or "Export")[:31]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    wrap_top = Alignment(wrap_text=True, vertical="top")

    for col_idx, (_, label) in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row in enumerate(rows, start=2):
        max_lines = 1
        for col_idx, (key, _) in enumerate(columns, start=1):
            multiline = key == "changes"
            value = _cell_value(row, key, multiline=multiline)
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            if multiline:
                cell.alignment = wrap_top
                max_lines = max(max_lines, value.count("\n") + 1)
        if max_lines > 1:
            sheet.row_dimensions[row_idx].height = min(15 * max_lines, 150)

    sample_rows = rows[:200]  # enough to size columns sensibly without scanning huge exports
    for col_idx, (key, label) in enumerate(columns, start=1):
        column_letter = sheet.cell(row=1, column=col_idx).column_letter
        if key == "changes":
            sheet.column_dimensions[column_letter].width = 50
            continue
        longest = max([len(label)] + [len(_cell_value(row, key)) for row in sample_rows])
        sheet.column_dimensions[column_letter].width = min(max(longest + 2, 10), 40)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# --- PDF ---------------------------------------------------------------------

_PDF_PAGE_WIDTH, _PDF_PAGE_HEIGHT = fitz.paper_size("a4-l")  # landscape fits more columns
_PDF_MARGIN = 28
_PDF_ROW_HEIGHT = 27  # tall enough for 2 wrapped lines instead of a single truncated one
_PDF_ACCOUNTS_ROW_HEIGHT = 32  # accounts rows carry a 2-line balance + date cell
_PDF_HEADER_HEIGHT = 22
_PDF_FONT_SIZE = 8
_PDF_TITLE_FONT_SIZE = 20
_PDF_HEADER_COLOR = (0.15, 0.39, 0.92)
_PDF_STRIPE_COLOR = (0.96, 0.97, 0.99)
_PDF_CHANGES_FILL = (0.93, 0.96, 1.0)
_PDF_KV_LABEL_WIDTH = 105
_PDF_KV_LINE_HEIGHT = 16
_PDF_KV_FONT_SIZE = 10
_PDF_COVER_BOX_BORDER = (0.82, 0.86, 0.94)
_PDF_COVER_BOX_FILL = (0.97, 0.98, 1.0)


class _PdfReport:
    """Owns page creation / the running y-cursor so each report type only
    has to worry about drawing its own rows. Page 1 gets a cover block
    (who the report is for, its date range/as-of, and any active filters);
    later pages just repeat the title and row count."""

    def __init__(self, doc, title: str, row_count: int, org: Organization, date_label: str, date_line: str | None, filter_lines: list[str]):
        self.doc = doc
        self.title = title
        self.row_count = row_count
        self.org = org
        self.date_label = date_label
        self.date_line = date_line
        self.filter_lines = filter_lines
        self._is_first_page = True
        self.page = None
        self.y = 0
        self.new_page()

    def new_page(self):
        page = self.doc.new_page(width=_PDF_PAGE_WIDTH, height=_PDF_PAGE_HEIGHT)
        usable_width = _PDF_PAGE_WIDTH - 2 * _PDF_MARGIN
        y = _PDF_MARGIN
        page.insert_text((_PDF_MARGIN, y + 18), self.title, fontsize=_PDF_TITLE_FONT_SIZE, fontname="hebo", color=(0.05, 0.05, 0.08))
        y += 18 + 14

        if self._is_first_page:
            kv_rows = []
            if self.org:
                kv_rows.append(("Name", self.org.name or "-"))
                kv_rows.append(("Email", self.org.email or "-"))
            if self.date_line:
                kv_rows.append((self.date_label, self.date_line))
            if self.filter_lines:
                kv_rows.append(("Filters", "; ".join(self.filter_lines)))

            if kv_rows:
                box_padding = 8
                box_height = len(kv_rows) * _PDF_KV_LINE_HEIGHT + box_padding * 2
                box_rect = fitz.Rect(_PDF_MARGIN, y, _PDF_MARGIN + usable_width, y + box_height)
                page.draw_rect(box_rect, color=_PDF_COVER_BOX_BORDER, fill=_PDF_COVER_BOX_FILL, width=0.75)

                row_y = y + box_padding
                value_max_x = _PDF_MARGIN + usable_width - box_padding
                for label, value in kv_rows:
                    label_rect = fitz.Rect(_PDF_MARGIN + box_padding, row_y, _PDF_MARGIN + box_padding + _PDF_KV_LABEL_WIDTH, row_y + _PDF_KV_LINE_HEIGHT)
                    _draw_cell(page, label_rect, label, fontsize=_PDF_KV_FONT_SIZE, fontname="hebo", color=(0.32, 0.34, 0.42))
                    value_rect = fitz.Rect(_PDF_MARGIN + box_padding + _PDF_KV_LABEL_WIDTH, row_y, value_max_x, row_y + _PDF_KV_LINE_HEIGHT)
                    _draw_cell(page, value_rect, value, fontsize=_PDF_KV_FONT_SIZE, color=(0.08, 0.08, 0.12))
                    row_y += _PDF_KV_LINE_HEIGHT

                y += box_height + 10

        generated_line = f"Generated {datetime.utcnow().strftime('%d %b %Y, %H:%M UTC')} - {self.row_count} rows"
        page.insert_text((_PDF_MARGIN, y + 8), generated_line, fontsize=8, fontname="helv", color=(0.4, 0.4, 0.4))
        y += 8 + 14

        self.page = page
        self.y = y
        self._is_first_page = False
        return page, y

    def ensure_space(self, needed_height: float, redraw_header=None):
        if self.y + needed_height > _PDF_PAGE_HEIGHT - _PDF_MARGIN:
            self.new_page()
            if redraw_header:
                self.y = redraw_header(self.page, self.y)


def _fit_text(text: str, max_width: float, fontsize: float = _PDF_FONT_SIZE, fontname: str = "helv") -> str:
    """Truncates text with '...' so it's guaranteed to fit in one line at
    max_width. insert_textbox()'s auto-wrap only breaks on spaces, so a
    single unbroken token wider than the box (an account number, a run-on
    narration, an ISO timestamp) gets silently dropped instead of drawn -
    that's what caused whole columns to vanish. Pre-fitting the text avoids
    relying on that wrap behavior at all."""
    if not text:
        return text
    if fitz.get_text_length(text, fontname=fontname, fontsize=fontsize) <= max_width:
        return text

    ellipsis = "..."
    if fitz.get_text_length(ellipsis, fontname=fontname, fontsize=fontsize) > max_width:
        return ""

    low, high = 0, len(text)
    fitted = ellipsis
    while low <= high:
        mid = (low + high) // 2
        candidate = text[:mid].rstrip() + ellipsis
        if fitz.get_text_length(candidate, fontname=fontname, fontsize=fontsize) <= max_width:
            fitted = candidate
            low = mid + 1
        else:
            high = mid - 1
    return fitted


def _draw_cell(page, rect: fitz.Rect, text: str, fontsize: float = _PDF_FONT_SIZE, color=(0.1, 0.1, 0.1), fontname: str = "helv"):
    """Draws single-line, guaranteed-to-fit cell text via insert_text
    (a fixed baseline point) rather than insert_textbox's word-wrap, which
    is what silently dropped long single-token cells."""
    fitted = _fit_text(text, rect.width, fontsize=fontsize, fontname=fontname)
    if not fitted:
        return
    baseline_y = min(rect.y0 + fontsize + 1, rect.y1)
    page.insert_text((rect.x0, baseline_y), fitted, fontsize=fontsize, fontname=fontname, color=color)


def _wrap_lines(text: str, max_width: float, max_lines: int = 2, fontsize: float = _PDF_FONT_SIZE, fontname: str = "helv") -> list[str]:
    """Word-wraps text across up to max_lines, each line individually
    guaranteed to fit max_width (falling back to _fit_text's char-level
    truncation for a single word too wide for one line on its own, and for
    whatever's left over once max_lines is used up)."""
    if not text:
        return []
    if fitz.get_text_length(text, fontname=fontname, fontsize=fontsize) <= max_width:
        return [text]

    def width(s):
        return fitz.get_text_length(s, fontname=fontname, fontsize=fontsize)

    words = text.split(" ")
    lines = []

    while words and len(lines) < max_lines:
        line = ""
        while words and width(f"{line} {words[0]}".strip()) <= max_width:
            line = f"{line} {words[0]}".strip()
            words.pop(0)

        if not line:
            # a single word doesn't fit even an empty line - truncate it
            line = _fit_text(words.pop(0), max_width, fontsize=fontsize, fontname=fontname)

        lines.append(line)

    if words and lines:
        # more text than max_lines can hold - mark the last line as cut off
        lines[-1] = _fit_text(f"{lines[-1]} {' '.join(words)}", max_width, fontsize=fontsize, fontname=fontname)

    return lines


def _draw_wrapped_cell(page, rect: fitz.Rect, text: str, max_lines: int = 2, fontsize: float = _PDF_FONT_SIZE, color=(0.1, 0.1, 0.1), fontname: str = "helv"):
    """Draws cell text across up to max_lines instead of truncating to one -
    each line is pre-fit the same guaranteed-safe way as _draw_cell."""
    lines = _wrap_lines(text, rect.width, max_lines=max_lines, fontsize=fontsize, fontname=fontname)
    line_height = fontsize + 3
    y = rect.y0 + fontsize + 1
    for line in lines:
        if y > rect.y1:
            break
        page.insert_text((rect.x0, y), line, fontsize=fontsize, fontname=fontname, color=color)
        y += line_height


# Relative width weights so columns with naturally short content (Type,
# Source) don't take the same room as ones that regularly run long
# (Narration, Counterparty). Any column key not listed here gets weight 1.0.
COLUMN_WIDTH_WEIGHTS = {
    "bank_name": 1.15,
    "account_holder_name": 1.3,
    "account_number": 1.2,
    "counterparty": 1.6,
    "txn_type": 0.55,
    "amount": 0.85,
    "mode": 0.75,
    "narration": 2.0,
    "source": 0.55,
    "changed_by": 0.85,
    "created_at": 1.15,
    "txn_narration": 2.0,
    "reason": 1.3,
    "ip_address": 0.9,
    "txn_amount": 0.75,
    "txn_id": 1.3,
    "account_type": 0.75,
    "delta": 0.85,
    "category": 0.9,
}
DEFAULT_COLUMN_WEIGHT = 1.0


def _weighted_col_widths(columns: list[tuple[str, str]], usable_width: float) -> list[float]:
    weights = [COLUMN_WIDTH_WEIGHTS.get(key, DEFAULT_COLUMN_WEIGHT) for key, _ in columns]
    total_weight = sum(weights) or len(columns)
    return [usable_width * (weight / total_weight) for weight in weights]


def _draw_table_header(page, y, columns, col_widths: list[float]):
    header_rect = fitz.Rect(_PDF_MARGIN, y, _PDF_MARGIN + sum(col_widths), y + _PDF_HEADER_HEIGHT)
    page.draw_rect(header_rect, color=None, fill=_PDF_HEADER_COLOR)
    x = _PDF_MARGIN
    for (_, label), width in zip(columns, col_widths):
        cell_rect = fitz.Rect(x + 3, y + 3, x + width - 3, y + _PDF_HEADER_HEIGHT - 3)
        _draw_cell(page, cell_rect, label, fontname="hebo", color=(1, 1, 1))
        x += width
    return y + _PDF_HEADER_HEIGHT


def _build_pdf(
    source: str,
    rows: list[dict],
    columns: list[tuple[str, str]],
    title: str,
    org: Organization,
    date_label: str,
    date_line: str | None,
    filter_lines: list[str],
) -> bytes:
    doc = fitz.open()
    report_args = (doc, rows, title, org, date_label, date_line, filter_lines)

    if source == "accounts":
        return _draw_accounts_pdf(*report_args, columns)
    if source == "audit-log":
        return _draw_audit_log_pdf(*report_args, columns)
    return _draw_generic_pdf(*report_args, columns)


def _draw_generic_pdf(doc, rows, title, org, date_label, date_line, filter_lines, columns: list[tuple[str, str]]) -> bytes:
    usable_width = _PDF_PAGE_WIDTH - 2 * _PDF_MARGIN
    col_widths = _weighted_col_widths(columns, usable_width)

    def redraw_header(page, y):
        return _draw_table_header(page, y, columns, col_widths)

    report = _PdfReport(doc, title, len(rows), org, date_label, date_line, filter_lines)
    report.y = redraw_header(report.page, report.y)

    for row_idx, row in enumerate(rows):
        report.ensure_space(_PDF_ROW_HEIGHT, redraw_header)
        page, y = report.page, report.y

        if row_idx % 2 == 1:
            page.draw_rect(fitz.Rect(_PDF_MARGIN, y, _PDF_MARGIN + usable_width, y + _PDF_ROW_HEIGHT), color=None, fill=_PDF_STRIPE_COLOR)

        x = _PDF_MARGIN
        for (key, _), width in zip(columns, col_widths):
            cell_rect = fitz.Rect(x + 3, y + 2, x + width - 3, y + _PDF_ROW_HEIGHT - 2)
            _draw_wrapped_cell(page, cell_rect, _cell_value(row, key, formatters=PDF_COLUMN_FORMATTERS))
            x += width

        report.y += _PDF_ROW_HEIGHT

    if not rows:
        report.page.insert_text((_PDF_MARGIN, report.y + 20), "No data available for this export.", fontsize=10, fontname="helv", color=(0.5, 0.5, 0.5))

    return _finish(doc)


# Mirrors the Accounts page's own BalanceCell: the amount on top (bold,
# dark), the "as of" date underneath (small, muted) - instead of a separate
# "Last Updated" column that can't say which balance it belongs to.
_ACCOUNTS_PDF_COLUMNS = [
    ("account_holder_name", "Account Holder"),
    ("account_number", "Account Number"),
    ("bank_name", "Bank"),
    ("category", "Category"),
    ("account_type", "Type"),
    ("statement_balance", "Balance - Statement"),
    ("calculated_balance", "Balance - Calculated"),
    ("delta", "Delta"),
    ("source", "Source"),
]
_ACCOUNTS_BALANCE_DATE_KEYS = {
    "statement_balance": "statement_updated_at",
    "calculated_balance": "calculated_updated_at",
}


def _draw_accounts_pdf(doc, rows, title, org, date_label, date_line, filter_lines, selected_columns: list[tuple[str, str]]) -> bytes:
    # Balance columns always carry their "as of" date stacked underneath,
    # mirroring the Accounts page's own BalanceCell - that pairing isn't a
    # separately toggleable column in the PDF the way it is in CSV/Excel.
    selected_keys = {key for key, _ in selected_columns}
    columns = [col for col in _ACCOUNTS_PDF_COLUMNS if col[0] in selected_keys] or _ACCOUNTS_PDF_COLUMNS
    row_height = _PDF_ACCOUNTS_ROW_HEIGHT
    usable_width = _PDF_PAGE_WIDTH - 2 * _PDF_MARGIN
    col_widths = _weighted_col_widths(columns, usable_width)

    def redraw_header(page, y):
        return _draw_table_header(page, y, columns, col_widths)

    report = _PdfReport(doc, title, len(rows), org, date_label, date_line, filter_lines)
    report.y = redraw_header(report.page, report.y)

    for row_idx, row in enumerate(rows):
        report.ensure_space(row_height, redraw_header)
        page, y = report.page, report.y

        if row_idx % 2 == 1:
            page.draw_rect(fitz.Rect(_PDF_MARGIN, y, _PDF_MARGIN + usable_width, y + row_height), color=None, fill=_PDF_STRIPE_COLOR)

        x = _PDF_MARGIN
        for (key, _), width in zip(columns, col_widths):
            date_key = _ACCOUNTS_BALANCE_DATE_KEYS.get(key)

            if date_key:
                amount_rect = fitz.Rect(x + 3, y + 3, x + width - 3, y + row_height / 2)
                _draw_cell(page, amount_rect, _cell_value(row, key, formatters=PDF_COLUMN_FORMATTERS), color=(0.05, 0.05, 0.05))

                date_text = _cell_value(row, date_key, formatters=PDF_COLUMN_FORMATTERS)
                date_rect = fitz.Rect(x + 3, y + row_height / 2, x + width - 3, y + row_height - 3)
                _draw_cell(page, date_rect, date_text, fontsize=_PDF_FONT_SIZE - 1, color=(0.55, 0.55, 0.55))
            else:
                cell_rect = fitz.Rect(x + 3, y + 3, x + width - 3, y + row_height - 3)
                _draw_wrapped_cell(page, cell_rect, _cell_value(row, key, formatters=PDF_COLUMN_FORMATTERS))

            x += width

        report.y += row_height

    if not rows:
        report.page.insert_text((_PDF_MARGIN, report.y + 20), "No data available for this export.", fontsize=10, fontname="helv", color=(0.5, 0.5, 0.5))

    return _finish(doc)


# Audit log: the "changes" diff doesn't belong squeezed into one narrow
# column, so it's drawn as an expanded strip directly under its row -
# same "Field: old -> new" wording the Audit Log page's own expanded view
# uses, just laid out for print instead of a click-to-expand panel.
_AUDIT_LOG_PDF_COLUMNS = [
    ("changed_by", "Operator"),
    ("created_at", "When"),
    ("txn_narration", "Narration"),
    ("txn_amount", "Amount"),
    ("reason", "Reason"),
    ("ip_address", "IP Address"),
    ("txn_id", "Transaction ID"),
]
_AUDIT_CHANGE_LINE_HEIGHT = 11
_AUDIT_CHANGE_BLOCK_PADDING = 8


def _draw_audit_log_pdf(doc, rows, title, org, date_label, date_line, filter_lines, selected_columns: list[tuple[str, str]]) -> bytes:
    # "Changes" is always rendered as the expanded block below its row - it
    # isn't a toggleable column in the PDF the way it is in CSV/Excel, since
    # the diff needs room a table cell can't give it.
    selected_keys = {key for key, _ in selected_columns}
    columns = [col for col in _AUDIT_LOG_PDF_COLUMNS if col[0] in selected_keys] or _AUDIT_LOG_PDF_COLUMNS
    usable_width = _PDF_PAGE_WIDTH - 2 * _PDF_MARGIN
    col_widths = _weighted_col_widths(columns, usable_width)

    def redraw_header(page, y):
        return _draw_table_header(page, y, columns, col_widths)

    report = _PdfReport(doc, title, len(rows), org, date_label, date_line, filter_lines)
    report.y = redraw_header(report.page, report.y)

    for row_idx, row in enumerate(rows):
        change_lines = _format_changes_lines(row.get("changes"))
        change_block_height = (
            _AUDIT_CHANGE_BLOCK_PADDING * 2 + _AUDIT_CHANGE_LINE_HEIGHT * (len(change_lines) + 1)
            if change_lines
            else 0
        )
        total_height = _PDF_ROW_HEIGHT + change_block_height

        report.ensure_space(total_height, redraw_header)
        page, y = report.page, report.y

        if row_idx % 2 == 1:
            page.draw_rect(fitz.Rect(_PDF_MARGIN, y, _PDF_MARGIN + usable_width, y + _PDF_ROW_HEIGHT), color=None, fill=_PDF_STRIPE_COLOR)

        x = _PDF_MARGIN
        for (key, _), width in zip(columns, col_widths):
            cell_rect = fitz.Rect(x + 3, y + 2, x + width - 3, y + _PDF_ROW_HEIGHT - 2)
            _draw_wrapped_cell(page, cell_rect, _cell_value(row, key, formatters=PDF_COLUMN_FORMATTERS))
            x += width

        y += _PDF_ROW_HEIGHT

        if change_lines:
            block_rect = fitz.Rect(_PDF_MARGIN, y, _PDF_MARGIN + usable_width, y + change_block_height)
            page.draw_rect(block_rect, color=None, fill=_PDF_CHANGES_FILL)

            text_x1 = _PDF_MARGIN + usable_width - 10
            line_y = y + _AUDIT_CHANGE_BLOCK_PADDING
            heading_rect = fitz.Rect(_PDF_MARGIN + 10, line_y, text_x1, line_y + _AUDIT_CHANGE_LINE_HEIGHT)
            _draw_cell(page, heading_rect, "Changes:", color=(0.2, 0.25, 0.5))
            line_y += _AUDIT_CHANGE_LINE_HEIGHT

            for line in change_lines:
                line_rect = fitz.Rect(_PDF_MARGIN + 20, line_y, text_x1, line_y + _AUDIT_CHANGE_LINE_HEIGHT)
                _draw_cell(page, line_rect, f"- {line}", color=(0.25, 0.3, 0.55))
                line_y += _AUDIT_CHANGE_LINE_HEIGHT

        report.y = y + change_block_height

    if not rows:
        report.page.insert_text((_PDF_MARGIN, report.y + 20), "No data available for this export.", fontsize=10, fontname="helv", color=(0.5, 0.5, 0.5))

    return _finish(doc)


def _finish(doc) -> bytes:
    pdf_bytes = doc.tobytes()
    doc.close()
    return pdf_bytes
